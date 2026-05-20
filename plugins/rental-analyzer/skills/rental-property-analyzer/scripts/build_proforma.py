#!/usr/bin/env python3
"""
Rental Property Analyzer — pro forma builder.

Input : a normalized JSON model (see references/data-model.md).
Output: an editable .xlsx workbook (live formulas + native charts) and a
        polished .pdf report (BiggerPockets-style visuals).

Usage:
    python3 build_proforma.py MODEL.json [--outdir DIR] [--no-pdf] [--basename NAME]

The .xlsx always builds (needs openpyxl). The .pdf builds when matplotlib is
importable; otherwise the script prints a one-line note and ships the xlsx alone.

Single source of truth: every number is computed once in compute_model().
The workbook writes live formulas (so the model stays editable) and the PDF
draws the computed values, so both always agree.
"""
import argparse
import datetime as _dt
import json
import os
import re
import sys


# --------------------------------------------------------------------------
# Math model
# --------------------------------------------------------------------------
def _num(x, default=0.0):
    """Coerce to float, tolerating None / "" / "$1,234" / "5%" style inputs."""
    if x is None or x == "":
        return float(default)
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).strip().replace(",", "").replace("$", "")
    pct = s.endswith("%")
    s = s.rstrip("%")
    try:
        v = float(s)
    except ValueError:
        return float(default)
    return v / 100.0 if pct else v


def pmt(rate_monthly, n_months, principal):
    """Level monthly payment (positive number)."""
    if n_months <= 0:
        return 0.0
    if rate_monthly == 0:
        return principal / n_months
    return principal * rate_monthly / (1 - (1 + rate_monthly) ** (-n_months))


def amortization_balances(principal, annual_rate, years, io_months=0):
    """Return {month: ending_balance} plus the level payment, computed monthly."""
    r = annual_rate / 12.0
    n = int(round(years * 12))
    amort_n = max(n - io_months, 0)
    level = pmt(r, amort_n, principal)
    bal = principal
    out = {0: principal}
    for m in range(1, n + 1):
        interest = bal * r
        if m <= io_months:
            payment = interest          # interest-only period
        else:
            payment = level
        principal_paid = payment - interest
        bal = max(bal - principal_paid, 0.0)
        out[m] = bal
    return out, level


def compute_model(model):
    """Turn the input JSON into a fully computed result dict used by both renderers."""
    a = model.get("assumptions", {})
    vac = _num(a.get("vacancy_rate", 0.05), 0.05)
    mgmt_pct = _num(a.get("management_pct", 0.08), 0.08)
    rm_pct = _num(a.get("repairs_pct", 0.05), 0.05)
    inc_g = _num(a.get("income_growth", 0.02), 0.02)
    exp_g = _num(a.get("expense_growth", 0.02), 0.02)
    appr = _num(a.get("appreciation", 0.03), 0.03)
    proj_years = a.get("projection_years", [1, 2, 5, 10, 15, 20, 30])

    prop = model.get("property", {})
    units = int(_num(prop.get("units", 1), 1)) or 1

    # ----- rent roll roll-ups -----
    roll = model.get("rent_roll", []) or []
    rr_current_mo = sum(_num(u.get("current_rent")) for u in roll)
    rr_market_mo = sum(_num(u.get("market_rent")) for u in roll)

    # ----- income -----
    inc = model.get("income", {})
    ri = inc.get("rental_income", {}) if isinstance(inc.get("rental_income"), dict) else {}
    rent_current = _num(ri.get("current")) or rr_current_mo * 12
    rent_proforma = _num(ri.get("proforma")) or rr_market_mo * 12

    vac_current = _num(inc.get("current_vacancy_loss"))          # negative or 0 (often embedded in T12 actuals)
    vac_proforma = -vac * rent_proforma

    other = []
    for o in (inc.get("other_income") or []):
        oc = _num(o.get("current"))
        op = _num(o.get("proforma")) if o.get("proforma") not in (None, "") else oc
        other.append({"label": o.get("label", "Other Income"), "current": oc, "proforma": op})
    other_cur = sum(o["current"] for o in other)
    other_pro = sum(o["proforma"] for o in other)

    egi_current = rent_current + vac_current + other_cur
    egi_proforma = rent_proforma + vac_proforma + other_pro

    # ----- expenses -----
    exp = model.get("expenses", {})

    def exp_pair(key, label, proforma_pct=None):
        e = exp.get(key, {}) if isinstance(exp.get(key), dict) else {}
        cur = _num(e.get("current"))
        if proforma_pct is not None:
            pro = proforma_pct * egi_proforma
        elif e.get("proforma") not in (None, ""):
            pro = _num(e.get("proforma"))
        else:
            pro = cur                                   # held flat at actuals
        return {"key": key, "label": label, "current": cur, "proforma": pro}

    expense_lines = [
        exp_pair("management", "Property Management", proforma_pct=mgmt_pct),
        exp_pair("property_taxes", "Property Taxes"),
        exp_pair("insurance", "Property Insurance"),
        exp_pair("water_sewer", "Water / Sewer"),
        exp_pair("electric", "Electricity"),
        exp_pair("gas", "Gas"),
        exp_pair("trash", "Trash / Waste"),
        exp_pair("repairs_maintenance", "Repairs & Maintenance", proforma_pct=rm_pct),
        exp_pair("payroll", "Payroll / On-site Labor"),
        exp_pair("landscaping", "Landscaping / Grounds"),
        exp_pair("make_ready", "Make Ready / Turnover"),
        exp_pair("general_admin", "General & Administrative"),
        exp_pair("reserves", "Replacement Reserves"),
    ]
    for o in (exp.get("other") or []):
        oc = _num(o.get("current"))
        op = _num(o.get("proforma")) if o.get("proforma") not in (None, "") else oc
        expense_lines.append({"key": None, "label": o.get("label", "Other Expense"),
                              "current": oc, "proforma": op})

    # Drop lines that are zero in BOTH columns AND carry no proforma rule, to keep it clean.
    rule_keys = {"management", "repairs_maintenance"}
    expense_lines = [e for e in expense_lines
                     if e["current"] or e["proforma"] or e["key"] in rule_keys]

    opex_current = sum(e["current"] for e in expense_lines)
    opex_proforma = sum(e["proforma"] for e in expense_lines)

    noi_current = egi_current - opex_current
    noi_proforma = egi_proforma - opex_proforma

    # ----- financing / returns -----
    fin = model.get("financing", {})
    price = _num(fin.get("purchase_price"))
    closing = _num(fin.get("closing_costs"))
    repairs = _num(fin.get("repair_costs"))
    fees = _num(fin.get("loan_fees"))
    ltv = _num(fin.get("ltv", 0.75), 0.75)
    rate = _num(fin.get("interest_rate", 0.065), 0.065)
    years = _num(fin.get("amortization_years", 30), 30)
    io_months = int(_num(fin.get("interest_only_months", 0)))

    loan = _num(fin.get("loan_amount")) or price * ltv
    down = price - loan
    cash_needed = down + closing + repairs + fees

    balances, level_pmt = amortization_balances(loan, rate, years, io_months)
    annual_ds = (loan * rate) if io_months >= 12 else level_pmt * 12
    monthly_pi = level_pmt

    def returns(noi):
        cap = noi / price if price else 0.0
        dscr = noi / annual_ds if annual_ds else 0.0
        fcf_yr = noi - annual_ds
        coc = fcf_yr / cash_needed if cash_needed else 0.0
        return {"cap_rate": cap, "dscr": dscr, "fcf_year": fcf_yr,
                "fcf_month": fcf_yr / 12, "fcf_unit_month": fcf_yr / 12 / units,
                "cash_on_cash": coc}

    ret_current = returns(noi_current)
    ret_proforma = returns(noi_proforma)
    grm = price / rent_proforma if rent_proforma else 0.0

    # ----- multi-year projection (base = proforma) -----
    projection = []
    for n in proj_years:
        income_n = egi_proforma * (1 + inc_g) ** (n - 1)
        expense_n = opex_proforma * (1 + exp_g) ** (n - 1)
        noi_n = income_n - expense_n
        cf_n = noi_n - annual_ds
        value_n = price * (1 + appr) ** n
        bal_n = balances.get(int(round(n * 12)), 0.0)
        equity_n = value_n - bal_n
        projection.append({
            "year": n, "income": income_n, "expenses": expense_n, "noi": noi_n,
            "cash_flow": cf_n, "value": value_n, "loan_balance": bal_n,
            "equity": equity_n,
            "coc": (cf_n / cash_needed if cash_needed else 0.0),
        })

    return {
        "property": {"address": prop.get("address", "Rental Property"),
                     "name": prop.get("name", ""),
                     "type": prop.get("property_type", ""),
                     "units": units,
                     "date": prop.get("analysis_date") or _dt.date.today().isoformat()},
        "assumptions": {"vacancy": vac, "mgmt_pct": mgmt_pct, "rm_pct": rm_pct,
                        "income_growth": inc_g, "expense_growth": exp_g,
                        "appreciation": appr, "projection_years": proj_years},
        "rent_roll": roll, "rr_current_mo": rr_current_mo, "rr_market_mo": rr_market_mo,
        "income": {"rent_current": rent_current, "rent_proforma": rent_proforma,
                   "vac_current": vac_current, "vac_proforma": vac_proforma,
                   "other": other, "egi_current": egi_current, "egi_proforma": egi_proforma},
        "expense_lines": expense_lines, "opex_current": opex_current, "opex_proforma": opex_proforma,
        "noi_current": noi_current, "noi_proforma": noi_proforma,
        "financing": {"price": price, "closing": closing, "repairs": repairs, "fees": fees,
                      "ltv": ltv, "rate": rate, "years": years, "io_months": io_months,
                      "loan": loan, "down": down, "cash_needed": cash_needed,
                      "monthly_pi": monthly_pi, "annual_ds": annual_ds},
        "ret_current": ret_current, "ret_proforma": ret_proforma, "grm": grm,
        "price_per_door": (price / units if units else 0.0),
        "projection": projection,
        "notes": model.get("notes", {}),
        "completeness": model.get("data_completeness", {"tier": 1, "missing": [], "flags": []}),
        "t12": model.get("t12", {}),
    }


# --------------------------------------------------------------------------
# Excel workbook (live formulas + native charts)
# --------------------------------------------------------------------------
ACCENT = "1F4E46"      # deep green
ACCENT2 = "3E7D6A"
LIGHT = "EAF1EE"
GREY = "6B6B6B"
LINE = "D9D9D9"
WHITE = "FFFFFF"
USD = '$#,##0'
USD2 = '$#,##0.00'
PCT = '0.0%'
NUM = '#,##0'


def build_workbook(M, path):
    import openpyxl
    from openpyxl.chart import PieChart, LineChart, Reference
    from openpyxl.chart.label import DataLabelList
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    wb.calculation.fullCalcOnLoad = True     # force Excel/Sheets/LibreOffice to recompute on open

    thin = Side(style="thin", color=LINE)
    box = Border(bottom=thin)

    def style(cell, *, bold=False, size=11, color="141414", fill=None,
              align=None, fmt=None, italic=False):
        cell.font = Font(bold=bold, size=size, color=color, italic=italic)
        if fill:
            cell.fill = PatternFill("solid", fgColor=fill)
        if align:
            cell.alignment = Alignment(horizontal=align, vertical="center", wrap_text=(align == "wrap"))
        if fmt:
            cell.number_format = fmt
        return cell

    # ===================== PROFORMA sheet =====================
    pf = wb.active
    pf.title = "Proforma"
    pf.sheet_view.showGridLines = False
    pf.column_dimensions["A"].width = 2
    pf.column_dimensions["B"].width = 30
    pf.column_dimensions["C"].width = 15
    pf.column_dimensions["D"].width = 15
    pf.column_dimensions["E"].width = 2
    pf.column_dimensions["F"].width = 42

    R = {}
    p = M["property"]
    style(pf["B1"], bold=True, size=16, color=WHITE, fill=ACCENT)
    pf["B1"] = p["address"]
    for c in ("C1", "D1", "E1", "F1"):
        pf[c].fill = PatternFill("solid", fgColor=ACCENT)
    pf.merge_cells("B1:F1")
    sub = f'{p["type"]} · {p["units"]} unit(s) · Analysis {p["date"]}'.strip(" ·")
    style(pf["B2"], size=9, color=GREY)
    pf["B2"] = sub

    style(pf["C4"], bold=True, align="center", color=ACCENT)
    pf["C4"] = "Current"
    style(pf["D4"], bold=True, align="center", color=ACCENT)
    pf["D4"] = "Proforma"

    row = 5

    def section(title):
        nonlocal row
        style(pf.cell(row=row, column=2, value=title), bold=True, size=11, color=WHITE, fill=ACCENT2)
        for col in (3, 4, 6):
            pf.cell(row=row, column=col).fill = PatternFill("solid", fgColor=ACCENT2)
        row += 1

    def line(label, cur=None, pro=None, note=None, *, bold=False, fmt=USD, fill=None):
        nonlocal row
        style(pf.cell(row=row, column=2, value=label), bold=bold, fill=fill)
        if cur is not None:
            style(pf.cell(row=row, column=3, value=cur), bold=bold, fmt=fmt, align="right", fill=fill)
        if pro is not None:
            style(pf.cell(row=row, column=4, value=pro), bold=bold, fmt=fmt, align="right", fill=fill)
        if note:
            style(pf.cell(row=row, column=6, value=note), size=9, color=GREY, italic=True)
        r = row
        row += 1
        return r

    inc = M["income"]
    notes = M["notes"]
    section("INCOME")
    R["rent"] = line("Rental Income", inc["rent_current"], inc["rent_proforma"],
                     notes.get("rental_income", "Current = in-place / T12 actual; Proforma = market rent × 12"))
    vac_note = f'Proforma vacancy @ {M["assumptions"]["vacancy"]*100:.0f}%; current reflected in actuals'
    R["vac"] = line("Vacancy & Credit Loss", inc["vac_current"], inc["vac_proforma"], vac_note)
    R["other_first"] = row
    for o in inc["other"]:
        line(o["label"], o["current"], o["proforma"])
    R["other_last"] = row - 1
    # EGI = rent + vacancy + other income block
    egi_cur_f = f'=C{R["rent"]}+C{R["vac"]}'
    egi_pro_f = f'=D{R["rent"]}+D{R["vac"]}'
    if inc["other"]:
        egi_cur_f += f'+SUM(C{R["other_first"]}:C{R["other_last"]})'
        egi_pro_f += f'+SUM(D{R["other_first"]}:D{R["other_last"]})'
    R["egi"] = line("Effective Gross Income (EGI)", egi_cur_f, egi_pro_f, bold=True, fill=LIGHT)

    row += 1
    section("OPERATING EXPENSES")
    R["exp_first"] = row
    for e in M["expense_lines"]:
        note = notes.get(e["key"]) if e["key"] else None
        if e["key"] == "management":
            pro = f'=D{R["egi"]}*{M["assumptions"]["mgmt_pct"]}'
            note = note or f'{M["assumptions"]["mgmt_pct"]*100:.0f}% of EGI (proforma); actual (current)'
        elif e["key"] == "repairs_maintenance":
            pro = f'=D{R["egi"]}*{M["assumptions"]["rm_pct"]}'
            note = note or f'{M["assumptions"]["rm_pct"]*100:.0f}% of EGI (proforma); actual (current)'
        else:
            pro = e["proforma"]
        line(e["label"], e["current"], pro, note)
    R["exp_last"] = row - 1
    R["opex"] = line("Total Operating Expenses",
                     f'=SUM(C{R["exp_first"]}:C{R["exp_last"]})',
                     f'=SUM(D{R["exp_first"]}:D{R["exp_last"]})', bold=True, fill=LIGHT)

    row += 1
    section("NET OPERATING INCOME")
    R["noi"] = line("NOI", f'=C{R["egi"]}-C{R["opex"]}', f'=D{R["egi"]}-D{R["opex"]}',
                    bold=True, fill=LIGHT)

    fin = M["financing"]
    row += 1
    section("ACQUISITION & FINANCING")
    R["price"] = line("Purchase Price", fin["price"], f'=C{row}')  # D mirrors C
    R["ppd"] = line("Price per Door", f'=C{R["price"]}/{p["units"]}', f'=D{R["price"]}/{p["units"]}')
    R["ltv"] = line("LTV", fin["ltv"], f'=C{row}', fmt=PCT)
    R["loan"] = line("Loan Amount", f'=C{R["price"]}*C{R["ltv"]}', f'=D{R["price"]}*D{R["ltv"]}',
                     "Or override on the Loan sheet")
    R["down"] = line("Down Payment", f'=C{R["price"]}-C{R["loan"]}', f'=D{R["price"]}-D{R["loan"]}')
    R["cash"] = line("Total Cash Needed", fin["cash_needed"], f'=C{row}',
                     "Down payment + closing + rehab + loan fees")
    R["ds"] = line("Annual Debt Service", "=Loan!B10", "=Loan!B10",
                   f'{fin["rate"]*100:.3f}% · {int(fin["years"])}-yr amortization (Loan sheet)')

    row += 1
    section("RETURNS")
    R["cap"] = line("Cap Rate (on price)", f'=C{R["noi"]}/C{R["price"]}',
                    f'=D{R["noi"]}/D{R["price"]}', fmt=PCT)
    R["dscr"] = line("DSCR", f'=C{R["noi"]}/C{R["ds"]}', f'=D{R["noi"]}/D{R["ds"]}', fmt='0.00')
    R["fcf_y"] = line("Free Cash Flow (Year)", f'=C{R["noi"]}-C{R["ds"]}', f'=D{R["noi"]}-D{R["ds"]}')
    R["fcf_m"] = line("Free Cash Flow (Month)", f'=C{R["fcf_y"]}/12', f'=D{R["fcf_y"]}/12')
    R["fcf_u"] = line("Cash Flow / Unit / Month", f'=C{R["fcf_m"]}/{p["units"]}',
                      f'=D{R["fcf_m"]}/{p["units"]}', fmt=USD2)
    R["coc"] = line("Cash-on-Cash Return", f'=C{R["fcf_y"]}/C{R["cash"]}',
                    f'=D{R["fcf_y"]}/D{R["cash"]}', fmt=PCT)
    R["grm"] = line("Gross Rent Multiplier", f'=C{R["price"]}/C{R["rent"]}',
                    f'=D{R["price"]}/D{R["rent"]}', fmt='0.0')

    pf.print_area = f"A1:F{row}"
    pf.page_setup.orientation = "portrait"
    pf.page_setup.fitToWidth = 1
    pf.page_setup.fitToHeight = 1
    pf.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)
    M["_R"] = R  # expose proforma row map for the Summary sheet

    # ===================== RENT ROLL sheet =====================
    rr = wb.create_sheet("Rent Roll")
    rr.sheet_view.showGridLines = False
    headers = ["Unit", "Type", "Sq Ft", "Current Rent", "Market Rent", "Status", "Notes"]
    widths = [12, 16, 9, 14, 14, 12, 28]
    for i, (h, w) in enumerate(zip(headers, widths), start=1):
        style(rr.cell(row=1, column=i, value=h), bold=True, color=WHITE, fill=ACCENT)
        rr.column_dimensions[get_column_letter(i)].width = w
    r = 2
    for u in M["rent_roll"]:
        rr.cell(row=r, column=1, value=u.get("unit", ""))
        rr.cell(row=r, column=2, value=u.get("type", ""))
        style(rr.cell(row=r, column=3, value=_num(u.get("sqft")) or None), fmt=NUM, align="right")
        style(rr.cell(row=r, column=4, value=_num(u.get("current_rent")) or None), fmt=USD, align="right")
        style(rr.cell(row=r, column=5, value=_num(u.get("market_rent")) or None), fmt=USD, align="right")
        rr.cell(row=r, column=6, value=u.get("status", ""))
        style(rr.cell(row=r, column=7, value=u.get("notes", "")), size=9, color=GREY, italic=True)
        r += 1
    last = r - 1
    style(rr.cell(row=r, column=2, value="TOTAL (monthly)"), bold=True, fill=LIGHT)
    style(rr.cell(row=r, column=4, value=f"=SUM(D2:D{last})" if last >= 2 else 0), bold=True, fmt=USD, align="right", fill=LIGHT)
    style(rr.cell(row=r, column=5, value=f"=SUM(E2:E{last})" if last >= 2 else 0), bold=True, fmt=USD, align="right", fill=LIGHT)
    for col in (1, 3, 6, 7):
        rr.cell(row=r, column=col).fill = PatternFill("solid", fgColor=LIGHT)

    # ===================== T12 sheet =====================
    t12 = M.get("t12") or {}
    if t12.get("rows"):
        ts = wb.create_sheet("T12")
        ts.sheet_view.showGridLines = False
        style(ts.cell(row=1, column=1, value=t12.get("period", "Trailing 12 Months")), bold=True, color=WHITE, fill=ACCENT)
        ts.cell(row=1, column=2).fill = PatternFill("solid", fgColor=ACCENT)
        ts.column_dimensions["A"].width = 38
        ts.column_dimensions["B"].width = 16
        style(ts.cell(row=2, column=1, value="Line Item"), bold=True)
        style(ts.cell(row=2, column=2, value="Annual (T12)"), bold=True, align="right")
        rr2 = 3
        for item in t12["rows"]:
            label = item.get("label", "")
            amt = _num(item.get("annual"))
            is_total = item.get("total", False)
            style(ts.cell(row=rr2, column=1, value=label), bold=is_total, fill=(LIGHT if is_total else None))
            style(ts.cell(row=rr2, column=2, value=amt), bold=is_total, fmt=USD, align="right",
                  fill=(LIGHT if is_total else None))
            rr2 += 1

    # ===================== LOAN sheet =====================
    ln = wb.create_sheet("Loan")
    ln.sheet_view.showGridLines = False
    ln.column_dimensions["A"].width = 26
    ln.column_dimensions["B"].width = 16
    style(ln.cell(row=1, column=1, value="Loan Amortization"), bold=True, size=14, color=WHITE, fill=ACCENT)
    ln.cell(row=1, column=2).fill = PatternFill("solid", fgColor=ACCENT)

    def lrow(rr_, label, value, fmt=None):
        style(ln.cell(row=rr_, column=1, value=label))
        c = ln.cell(row=rr_, column=2, value=value)
        style(c, align="right", fmt=fmt)
        return rr_

    lrow(3, "Loan Amount", "=Proforma!C{}".format(R["loan"]), USD)
    lrow(4, "Interest Rate", fin["rate"], PCT)
    lrow(5, "Amortization (years)", int(fin["years"]), NUM)
    lrow(6, "Interest-Only (months)", int(fin["io_months"]), NUM)
    style(ln.cell(row=7, column=1, value="Monthly Payment (P&I)"), bold=True)
    style(ln.cell(row=7, column=2, value="=-PMT(B4/12,(B5*12)-B6,B3)"), bold=True, align="right", fmt=USD2)
    style(ln.cell(row=8, column=1, value="Monthly Payment (interest-only)"))
    style(ln.cell(row=8, column=2, value="=B3*B4/12"), align="right", fmt=USD2)
    style(ln.cell(row=10, column=1, value="Annual Debt Service"), bold=True, fill=LIGHT)
    # interest-only if IO covers the whole first year, else level P&I × 12
    style(ln.cell(row=10, column=2, value="=IF(B6>=12,B8*12,B7*12)"), bold=True, align="right", fmt=USD, fill=LIGHT)

    # End-of-year balance table (years 1..10) using FV/cumulative — simple, transparent.
    style(ln.cell(row=12, column=1, value="End of Year"), bold=True)
    style(ln.cell(row=12, column=2, value="Loan Balance"), bold=True, align="right")
    base = M["financing"]["loan"]
    bal_map, _ = amortization_balances(base, fin["rate"], fin["years"], fin["io_months"])
    for yr in range(1, 11):
        rr3 = 12 + yr
        style(ln.cell(row=rr3, column=1, value=yr), align="right")
        style(ln.cell(row=rr3, column=2, value=round(bal_map.get(yr * 12, 0.0), 2)), align="right", fmt=USD)

    # ===================== ASSUMPTIONS sheet =====================
    asm = wb.create_sheet("Assumptions")
    asm.sheet_view.showGridLines = False
    asm.column_dimensions["A"].width = 30
    asm.column_dimensions["B"].width = 14
    asm.column_dimensions["C"].width = 50
    style(asm.cell(row=1, column=1, value="Editable Assumptions"), bold=True, size=14, color=WHITE, fill=ACCENT)
    for col in (2, 3):
        asm.cell(row=1, column=col).fill = PatternFill("solid", fgColor=ACCENT)
    A = M["assumptions"]
    asm_rows = [
        ("Vacancy rate", A["vacancy"], PCT, "Applied to gross market rent in the Proforma column"),
        ("Management % of EGI", A["mgmt_pct"], PCT, "Proforma management fee"),
        ("Repairs & Maint % of EGI", A["rm_pct"], PCT, "Proforma R&M reserve"),
        ("Income growth / yr", A["income_growth"], PCT, "Projection: annual revenue growth"),
        ("Expense growth / yr", A["expense_growth"], PCT, "Projection: annual expense growth"),
        ("Appreciation / yr", A["appreciation"], PCT, "Projection: property value growth"),
    ]
    rr4 = 3
    for label, val, fmt, note in asm_rows:
        style(asm.cell(row=rr4, column=1, value=label))
        style(asm.cell(row=rr4, column=2, value=val), align="right", fmt=fmt)
        style(asm.cell(row=rr4, column=3, value=note), size=9, color=GREY, italic=True)
        rr4 += 1

    # ===================== SUMMARY sheet (dashboard) =====================
    sm = wb.create_sheet("Summary")
    wb.move_sheet("Summary", -(len(wb.sheetnames) - 1))  # make it first
    sm.sheet_view.showGridLines = False
    for col, w in zip("ABCDEFGHIJKLM", [2, 18, 13, 13, 4, 18, 13, 13, 4, 13, 13, 13, 13]):
        sm.column_dimensions[col].width = w

    style(sm["B2"], bold=True, size=18, color=WHITE, fill=ACCENT)
    sm["B2"] = p["address"]
    for c in "CDEFGHIJKLM":
        sm[f"{c}2"].fill = PatternFill("solid", fgColor=ACCENT)
    sm.merge_cells("B2:M2")
    style(sm["B3"], size=10, color=GREY)
    sm["B3"] = sub
    sm.merge_cells("B3:M3")

    # KPI band: label row 5, Current row 6, Proforma row 7
    kpis = [
        ("NOI", f'Proforma!C{R["noi"]}', f'Proforma!D{R["noi"]}', USD),
        ("Cap Rate", f'Proforma!C{R["cap"]}', f'Proforma!D{R["cap"]}', PCT),
        ("DSCR", f'Proforma!C{R["dscr"]}', f'Proforma!D{R["dscr"]}', '0.00'),
        ("Cash-on-Cash", f'Proforma!C{R["coc"]}', f'Proforma!D{R["coc"]}', PCT),
        ("Cash Flow / Yr", f'Proforma!C{R["fcf_y"]}', f'Proforma!D{R["fcf_y"]}', USD),
        ("Price / Door", f'Proforma!C{R["ppd"]}', f'Proforma!D{R["ppd"]}', USD),
    ]
    col = 2
    style(sm.cell(row=5, column=1, value=""), size=9)
    style(sm.cell(row=6, column=1, value="Current"), size=9, bold=True, color=GREY, align="right")
    style(sm.cell(row=7, column=1, value="Proforma"), size=9, bold=True, color=ACCENT, align="right")
    for label, cur_ref, pro_ref, fmt in kpis:
        style(sm.cell(row=5, column=col, value=label), size=9, bold=True, color=GREY, fill=LIGHT, align="center")
        style(sm.cell(row=6, column=col, value=f'={cur_ref}'), size=11, color=GREY, fill=LIGHT, align="center", fmt=fmt)
        style(sm.cell(row=7, column=col, value=f'={pro_ref}'), size=14, bold=True, color=ACCENT, fill=LIGHT, align="center", fmt=fmt)
        col += 1

    # --- chart-source data (formula-linked to Proforma so charts stay live) ---
    # Expense composition (proforma column)
    sm["B40"] = "Expense"
    sm["C40"] = "Amount"
    er = 41
    for e in M["expense_lines"]:
        # find this expense's proforma cell on Proforma sheet
        sm.cell(row=er, column=2, value=e["label"])
        # reference proforma value by recomputing position: expenses occupy exp_first..exp_last in order
        idx = M["expense_lines"].index(e)
        pf_row = R["exp_first"] + idx
        sm.cell(row=er, column=3, value=f"=Proforma!D{pf_row}")
        er += 1
    exp_data_last = er - 1

    # Income composition (proforma): net rental + each other income
    sm["E40"] = "Income"
    sm["F40"] = "Amount"
    ir = 41
    sm.cell(row=ir, column=5, value="Net Rental")
    sm.cell(row=ir, column=6, value=f'=Proforma!D{R["rent"]}+Proforma!D{R["vac"]}')
    ir += 1
    for j, o in enumerate(M["income"]["other"]):
        sm.cell(row=ir, column=5, value=o["label"])
        sm.cell(row=ir, column=6, value=f'=Proforma!D{R["other_first"] + j}')
        ir += 1
    inc_data_last = ir - 1

    # Projection table (values; drives line charts)
    proj = M["projection"]
    headers = ["Year", "Income", "Expenses", "Cash Flow", "Property Value", "Loan Balance", "Equity"]
    for i, h in enumerate(headers):
        sm.cell(row=60, column=2 + i, value=h)
    for k, pr in enumerate(proj):
        rr5 = 61 + k
        sm.cell(row=rr5, column=2, value=pr["year"])
        sm.cell(row=rr5, column=3, value=round(pr["income"], 2))
        sm.cell(row=rr5, column=4, value=round(pr["expenses"], 2))
        sm.cell(row=rr5, column=5, value=round(pr["cash_flow"], 2))
        sm.cell(row=rr5, column=6, value=round(pr["value"], 2))
        sm.cell(row=rr5, column=7, value=round(pr["loan_balance"], 2))
        sm.cell(row=rr5, column=8, value=round(pr["equity"], 2))
    proj_last = 60 + len(proj)

    # --- charts ---
    pie_e = PieChart()
    pie_e.title = "Operating Expenses (Proforma)"
    pie_e.height = 7
    pie_e.width = 11
    data = Reference(sm, min_col=3, min_row=40, max_row=exp_data_last)
    cats = Reference(sm, min_col=2, min_row=41, max_row=exp_data_last)
    pie_e.add_data(data, titles_from_data=True)
    pie_e.set_categories(cats)
    pie_e.dataLabels = DataLabelList()
    pie_e.dataLabels.showPercent = True
    sm.add_chart(pie_e, "B9")

    pie_i = PieChart()
    pie_i.title = "Income (Proforma)"
    pie_i.height = 7
    pie_i.width = 11
    data = Reference(sm, min_col=6, min_row=40, max_row=inc_data_last)
    cats = Reference(sm, min_col=5, min_row=41, max_row=inc_data_last)
    pie_i.add_data(data, titles_from_data=True)
    pie_i.set_categories(cats)
    pie_i.dataLabels = DataLabelList()
    pie_i.dataLabels.showPercent = True
    sm.add_chart(pie_i, "F9")

    line1 = LineChart()
    line1.title = "Income · Expenses · Cash Flow"
    line1.height = 7
    line1.width = 11
    d = Reference(sm, min_col=3, max_col=5, min_row=60, max_row=proj_last)
    c = Reference(sm, min_col=2, min_row=61, max_row=proj_last)
    line1.add_data(d, titles_from_data=True)
    line1.set_categories(c)
    sm.add_chart(line1, "B25")

    line2 = LineChart()
    line2.title = "Equity · Loan Balance · Value"
    line2.height = 7
    line2.width = 11
    d = Reference(sm, min_col=6, max_col=8, min_row=60, max_row=proj_last)
    c = Reference(sm, min_col=2, min_row=61, max_row=proj_last)
    line2.add_data(d, titles_from_data=True)
    line2.set_categories(c)
    sm.add_chart(line2, "F25")

    # data-completeness banner (Tier 2/3)
    comp = M["completeness"]
    if comp.get("tier", 1) != 1 or comp.get("missing"):
        msg = "DATA COMPLETENESS: Tier {} — still needed: {}".format(
            comp.get("tier", "?"), ", ".join(comp.get("missing", [])) or "see notes")
        style(sm["B42"], bold=True, color="9A3B2E")
        sm["B42"] = msg

    sm.print_area = "B2:M38"
    sm.page_setup.orientation = "landscape"
    sm.page_setup.fitToWidth = 1
    sm.page_setup.fitToHeight = 1
    sm.sheet_properties.pageSetUpPr = openpyxl.worksheet.properties.PageSetupProperties(fitToPage=True)

    wb.save(path)
    return path


# --------------------------------------------------------------------------
# PDF report (matplotlib — BiggerPockets-style visuals)
# --------------------------------------------------------------------------
def _money(v):
    return ("-$" if v < 0 else "$") + "{:,.0f}".format(abs(v))


def build_pdf(M, path):
    import matplotlib
    matplotlib.use("Agg")
    import matplotlib.pyplot as plt
    from matplotlib.backends.backend_pdf import PdfPages
    from matplotlib.patches import FancyBboxPatch

    green = "#1F4E46"
    green2 = "#3E7D6A"
    light = "#EAF1EE"
    grey = "#6B6B6B"
    pie_colors = ["#1F4E46", "#3E7D6A", "#6FA292", "#A7C7BC", "#C9803A", "#D9A566",
                  "#8A8D91", "#B9BCC0", "#5B6E8C", "#9FB0C9", "#7A5C7B", "#B49AB5"]
    p = M["property"]

    with PdfPages(path) as pdf:
        # ---------- Page 1: dashboard ----------
        fig = plt.figure(figsize=(11, 8.5))
        fig.patch.set_facecolor("white")

        # header band
        hb = fig.add_axes([0, 0.92, 1, 0.08]); hb.axis("off")
        hb.add_patch(plt.Rectangle((0, 0), 1, 1, color=green))
        hb.text(0.04, 0.60, p["address"], color="white", fontsize=17, fontweight="bold", va="center")
        sub = " · ".join([x for x in [p["type"], f'{p["units"]} unit(s)', f'Analysis {p["date"]}'] if x and "unit" not in x or p["units"]])
        sub = " · ".join([x for x in [p["type"], f'{p["units"]} unit(s)', f'Analysis {p["date"]}'] if x])
        hb.text(0.04, 0.20, sub, color="#D5E3DD", fontsize=9, va="center")

        # KPI band (proforma headline, current as small sub-line)
        rc, rp = M["ret_current"], M["ret_proforma"]
        fin = M["financing"]
        kpis = [
            ("NOI", _money(M["noi_proforma"]), _money(M["noi_current"]) + " cur"),
            ("Cap Rate", f'{rp["cap_rate"]*100:.2f}%', f'{rc["cap_rate"]*100:.2f}% cur'),
            ("DSCR", f'{rp["dscr"]:.2f}', f'{rc["dscr"]:.2f} cur'),
            ("Cash-on-Cash", f'{rp["cash_on_cash"]*100:.2f}%', f'{rc["cash_on_cash"]*100:.2f}% cur'),
            ("Cash Flow/Yr", _money(rp["fcf_year"]), _money(rc["fcf_year"]) + " cur"),
            ("Price/Door", _money(M["price_per_door"]), f'{p["units"]} doors'),
        ]
        n = len(kpis)
        gap = 0.012
        w = (1 - 0.08 - (n - 1) * gap) / n
        for i, (lab, val, sub2) in enumerate(kpis):
            x = 0.04 + i * (w + gap)
            ax = fig.add_axes([x, 0.78, w, 0.11]); ax.axis("off")
            ax.add_patch(FancyBboxPatch((0, 0), 1, 1, boxstyle="round,pad=0.02",
                                        facecolor=light, edgecolor=green, linewidth=0.8))
            ax.text(0.5, 0.78, lab, ha="center", va="center", fontsize=8.5, color=grey)
            ax.text(0.5, 0.44, val, ha="center", va="center", fontsize=15, color=green, fontweight="bold")
            ax.text(0.5, 0.16, sub2, ha="center", va="center", fontsize=7, color=grey)

        # expense pie
        exp = [e for e in M["expense_lines"] if e["proforma"] > 0]
        axp = fig.add_axes([0.05, 0.43, 0.26, 0.30])
        if exp:
            axp.pie([e["proforma"] for e in exp], labels=[e["label"] for e in exp],
                    autopct="%1.0f%%", colors=pie_colors[:len(exp)],
                    textprops={"fontsize": 6.0}, pctdistance=0.75,
                    wedgeprops={"linewidth": 0.5, "edgecolor": "white"})
        axp.set_title("Operating Expenses (Proforma)", fontsize=9.5, color=green, fontweight="bold")

        # income pie
        inc = M["income"]
        islices = [("Net Rental", inc["rent_proforma"] + inc["vac_proforma"])] + \
                  [(o["label"], o["proforma"]) for o in inc["other"] if o["proforma"] > 0]
        axi = fig.add_axes([0.37, 0.43, 0.26, 0.30])
        axi.pie([s[1] for s in islices], labels=[s[0] for s in islices],
                autopct="%1.0f%%", colors=pie_colors[:len(islices)],
                textprops={"fontsize": 6.0}, pctdistance=0.75,
                wedgeprops={"linewidth": 0.5, "edgecolor": "white"})
        axi.set_title("Income (Proforma)", fontsize=9.5, color=green, fontweight="bold")

        # projection line chart (income/exp/cashflow)
        proj = M["projection"]
        yrs = [pr["year"] for pr in proj]
        axl = fig.add_axes([0.70, 0.46, 0.26, 0.27])
        axl.plot(yrs, [pr["income"] for pr in proj], marker="o", ms=3, color=green, label="Income")
        axl.plot(yrs, [pr["expenses"] for pr in proj], marker="o", ms=3, color="#C9803A", label="Expenses")
        axl.plot(yrs, [pr["cash_flow"] for pr in proj], marker="o", ms=3, color=green2, label="Cash Flow")
        axl.legend(fontsize=6.5, loc="upper left")
        axl.set_title("Projection ($/yr)", fontsize=9.5, color=green, fontweight="bold")
        axl.tick_params(labelsize=6.5)
        axl.yaxis.set_major_formatter(matplotlib.ticker.FuncFormatter(lambda v, _: f'{v/1000:.0f}k'))
        for s in ("top", "right"):
            axl.spines[s].set_visible(False)

        # Current vs Proforma table
        axt = fig.add_axes([0.05, 0.05, 0.58, 0.33]); axt.axis("off")
        rows = [
            ["Rental Income", _money(inc["rent_current"]), _money(inc["rent_proforma"])],
            ["Vacancy & Credit Loss", _money(inc["vac_current"]), _money(inc["vac_proforma"])],
        ]
        for o in inc["other"]:
            rows.append([o["label"], _money(o["current"]), _money(o["proforma"])])
        rows.append(["Effective Gross Income", _money(inc["egi_current"]), _money(inc["egi_proforma"])])
        rows.append(["Total Operating Expenses", _money(M["opex_current"]), _money(M["opex_proforma"])])
        rows.append(["Net Operating Income", _money(M["noi_current"]), _money(M["noi_proforma"])])
        rows.append(["Annual Debt Service", _money(fin["annual_ds"]), _money(fin["annual_ds"])])
        rows.append(["Free Cash Flow", _money(rc["fcf_year"]), _money(rp["fcf_year"])])
        tbl = axt.table(cellText=rows, colLabels=["Line Item", "Current", "Proforma"],
                        loc="upper center", cellLoc="left", colWidths=[0.5, 0.25, 0.25])
        tbl.auto_set_font_size(False)
        tbl.set_fontsize(8)
        tbl.scale(1, 1.35)
        bold_rows = {"Effective Gross Income", "Net Operating Income", "Free Cash Flow"}
        for (r, cc), cell in tbl.get_celld().items():
            cell.set_edgecolor("#E2E2E2")
            if r == 0:
                cell.set_facecolor(green); cell.set_text_props(color="white", fontweight="bold")
            elif rows[r - 1][0] in bold_rows:
                cell.set_facecolor(light); cell.set_text_props(fontweight="bold")

        # financing summary box
        axf = fig.add_axes([0.66, 0.05, 0.30, 0.33]); axf.axis("off")
        axf.text(0, 1.0, "Acquisition & Financing", fontsize=9.5, color=green, fontweight="bold", va="top")
        flines = [
            ("Purchase Price", _money(fin["price"])),
            ("Loan Amount", _money(fin["loan"]) + f'  ({fin["ltv"]*100:.0f}% LTV)'),
            ("Down Payment", _money(fin["down"])),
            ("Total Cash Needed", _money(fin["cash_needed"])),
            ("Interest Rate", f'{fin["rate"]*100:.3f}%'),
            ("Amortization", f'{int(fin["years"])} yrs'),
            ("Monthly P&I", _money(fin["monthly_pi"])),
            ("GRM", f'{M["grm"]:.1f}'),
        ]
        y = 0.86
        for lab, val in flines:
            axf.text(0.0, y, lab, fontsize=8, color=grey, va="top")
            axf.text(1.0, y, val, fontsize=8, color="#141414", va="top", ha="right", fontweight="bold")
            y -= 0.11

        # completeness note
        comp = M["completeness"]
        if comp.get("tier", 1) != 1 or comp.get("missing"):
            fig.text(0.05, 0.018, "Data completeness: Tier {} — still needed: {}".format(
                comp.get("tier", "?"), ", ".join(comp.get("missing", [])) or "see notes"),
                fontsize=7.5, color="#9A3B2E", style="italic")
        fig.text(0.96, 0.018, "Generated " + p["date"], fontsize=7, color=grey, ha="right")
        pdf.savefig(fig, facecolor="white")
        plt.close(fig)

        # ---------- Page 2: equity build + projection table ----------
        fig2 = plt.figure(figsize=(11, 8.5))
        hb = fig2.add_axes([0, 0.92, 1, 0.08]); hb.axis("off")
        hb.add_patch(plt.Rectangle((0, 0), 1, 1, color=green))
        hb.text(0.04, 0.5, "Long-Term Projection", color="white", fontsize=16, fontweight="bold", va="center")

        ax2 = fig2.add_axes([0.07, 0.52, 0.86, 0.33])
        ax2.plot(yrs, [pr["value"] for pr in proj], marker="o", ms=3, color=green, label="Property Value")
        ax2.plot(yrs, [pr["equity"] for pr in proj], marker="o", ms=3, color=green2, label="Equity")
        ax2.plot(yrs, [pr["loan_balance"] for pr in proj], marker="o", ms=3, color="#C9803A", label="Loan Balance")
        ax2.legend(fontsize=8)
        ax2.set_title("Equity · Loan Balance · Property Value", fontsize=10, color=green, fontweight="bold")
        ax2.yaxis.set_major_formatter(matplotlib.ticker.FuncFormatter(lambda v, _: f'${v/1000:.0f}k'))
        ax2.tick_params(labelsize=8)
        ax2.set_xlabel("Year", fontsize=8)
        for s in ("top", "right"):
            ax2.spines[s].set_visible(False)

        axt2 = fig2.add_axes([0.07, 0.07, 0.86, 0.38]); axt2.axis("off")
        cols = ["Year"] + [str(pr["year"]) for pr in proj]
        def fmt_row(label, key, money=True, pct=False):
            out = [label]
            for pr in proj:
                v = pr[key]
                out.append(_money(v) if money else (f'{v*100:.1f}%' if pct else f'{v:,.0f}'))
            return out
        trows = [
            fmt_row("Income", "income"),
            fmt_row("Expenses", "expenses"),
            fmt_row("NOI", "noi"),
            fmt_row("Cash Flow", "cash_flow"),
            fmt_row("Cash-on-Cash", "coc", money=False, pct=True),
            fmt_row("Property Value", "value"),
            fmt_row("Loan Balance", "loan_balance"),
            fmt_row("Equity", "equity"),
        ]
        tbl2 = axt2.table(cellText=trows, colLabels=cols, loc="upper center", cellLoc="center")
        tbl2.auto_set_font_size(False)
        tbl2.set_fontsize(7.5)
        tbl2.scale(1, 1.4)
        for (r, cc), cell in tbl2.get_celld().items():
            cell.set_edgecolor("#E2E2E2")
            if r == 0:
                cell.set_facecolor(green); cell.set_text_props(color="white", fontweight="bold")
            elif cc == 0:
                cell.set_text_props(fontweight="bold"); cell.set_facecolor(light)
        a = M["assumptions"]
        fig2.text(0.07, 0.03, "Assumptions: income +{:.0f}%/yr · expenses +{:.0f}%/yr · appreciation +{:.0f}%/yr · vacancy {:.0f}% · mgmt {:.0f}% · R&M {:.0f}%".format(
            a["income_growth"]*100, a["expense_growth"]*100, a["appreciation"]*100,
            a["vacancy"]*100, a["mgmt_pct"]*100, a["rm_pct"]*100), fontsize=7.5, color=grey, style="italic")
        pdf.savefig(fig2, facecolor="white")
        plt.close(fig2)
    return path


# --------------------------------------------------------------------------
def slugify(s):
    s = re.sub(r"[^A-Za-z0-9]+", "-", s or "").strip("-").lower()
    return s or "rental-property"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("model", help="Path to the normalized JSON model")
    ap.add_argument("--outdir", default=".")
    ap.add_argument("--basename", default=None)
    ap.add_argument("--no-pdf", action="store_true")
    args = ap.parse_args()

    with open(args.model) as f:
        raw = json.load(f)
    M = compute_model(raw)

    os.makedirs(args.outdir, exist_ok=True)
    base = args.basename or slugify(M["property"]["address"]) + "-proforma"
    xlsx_path = os.path.join(args.outdir, base + ".xlsx")
    build_workbook(M, xlsx_path)
    print("XLSX:", os.path.abspath(xlsx_path))

    if not args.no_pdf:
        try:
            pdf_path = os.path.join(args.outdir, base + ".pdf")
            build_pdf(M, pdf_path)
            print("PDF:", os.path.abspath(pdf_path))
        except ImportError:
            print("PDF: skipped (matplotlib not installed — `pip install matplotlib`, or open the "
                  "workbook's Summary sheet and Print → Save as PDF)")

    # machine-readable echo of headline numbers (handy for the chat summary)
    print("HEADLINE:", json.dumps({
        "noi_current": round(M["noi_current"], 2), "noi_proforma": round(M["noi_proforma"], 2),
        "cap_current": round(M["ret_current"]["cap_rate"], 4), "cap_proforma": round(M["ret_proforma"]["cap_rate"], 4),
        "dscr_proforma": round(M["ret_proforma"]["dscr"], 3),
        "coc_proforma": round(M["ret_proforma"]["cash_on_cash"], 4),
        "fcf_year_proforma": round(M["ret_proforma"]["fcf_year"], 2),
    }))


if __name__ == "__main__":
    main()
