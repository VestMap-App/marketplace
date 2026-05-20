#!/usr/bin/env python3
"""
Dump an .xlsx (rent roll / T12 / loan tab) to plain text so it can be read and
mapped into the pro forma model.

Usage:
    python3 dump_xlsx.py FILE.xlsx [--sheet NAME] [--max-rows N]

Prints every non-empty cell as `COORD=value`, one row per line, per sheet.
Needs openpyxl:  python3 -m pip install --quiet openpyxl
"""
import argparse
import sys


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("path")
    ap.add_argument("--sheet", default=None, help="Only dump this sheet")
    ap.add_argument("--max-rows", type=int, default=400)
    args = ap.parse_args()

    try:
        import openpyxl
    except ImportError:
        sys.exit("openpyxl not installed — run: python3 -m pip install --quiet openpyxl")

    wb_v = openpyxl.load_workbook(args.path, data_only=True)   # computed values
    wb_f = openpyxl.load_workbook(args.path, data_only=False)  # formulas
    print(f"FILE: {args.path}")
    print(f"SHEETS: {wb_v.sheetnames}")

    names = [args.sheet] if args.sheet else wb_v.sheetnames
    for name in names:
        if name not in wb_v.sheetnames:
            print(f"\n(skip — no sheet named {name!r})")
            continue
        ws, wsf = wb_v[name], wb_f[name]
        print(f"\n{'='*70}\nSHEET: {name!r}  dims={ws.dimensions}  rows={ws.max_row} cols={ws.max_column}\n{'='*70}")
        for i, row in enumerate(ws.iter_rows()):
            if i >= args.max_rows:
                print(f"... (truncated at {args.max_rows} rows; pass --max-rows to see more)")
                break
            cells = []
            for c in row:
                if c.value is None:
                    continue
                formula = wsf[c.coordinate].value
                if isinstance(formula, str) and formula.startswith("="):
                    cells.append(f"{c.coordinate}={c.value!r} [{formula}]")
                else:
                    cells.append(f"{c.coordinate}={c.value!r}")
            if cells:
                print(" | ".join(cells))


if __name__ == "__main__":
    main()
